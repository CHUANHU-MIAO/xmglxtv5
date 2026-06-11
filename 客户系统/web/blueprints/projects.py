import datetime
import io
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from web.models import Project, User, Log, Attachment
from web.extensions import db
from web.services.project_service import get_index_statistics

projects_bp = Blueprint('projects', __name__)

PROVINCES = [
    '北京市', '天津市', '上海市', '重庆市',
    '河北省', '山西省', '辽宁省', '吉林省', '黑龙江省',
    '江苏省', '浙江省', '安徽省', '福建省', '江西省', '山东省',
    '河南省', '湖北省', '湖南省', '广东省', '海南省',
    '四川省', '贵州省', '云南省', '陕西省', '甘肃省', '青海省',
    '内蒙古自治区', '广西壮族自治区', '西藏自治区', '宁夏回族自治区', '新疆维吾尔自治区',
    '香港特别行政区', '澳门特别行政区', '台湾省'
]


@projects_bp.route('/home')
@login_required
def index():
    page = request.args.get('page', 1, type=int)
    keyword = request.args.get('keyword', '', type=str)

    query = Project.query.filter_by(is_valid=1)
    if keyword:
        query = query.filter(Project.name.contains(keyword))

    projects = query.order_by(Project.updated_at.desc()).paginate(page=page, per_page=20, error_out=False)

    stats = get_index_statistics(months=12)

    return render_template('index.html',
                           projects=projects,
                           month_data=stats['month_data'],
                           month_labels=stats['month_labels'],
                           area_dict=stats['area_dict'],
                           eng_dict=stats['eng_dict'],
                           year_total=stats['year_total'],
                           page=page,
                           keyword=keyword,
                           role=current_user.role)


@projects_bp.route('/export_projects')
@login_required
def export_projects():
    keyword = request.args.get('keyword', '', type=str)
    query = Project.query.filter_by(is_valid=1)
    if keyword:
        query = query.filter(Project.name.contains(keyword))
    projects = query.order_by(Project.start_date.desc()).all()

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = '项目列表'

    headers = [
        '序号', '项目名称', '业主单位', '项目所在地', '总投资(万元)', '合同金额(万元)',
        '合同情况', '开票情况', '已开票金额(万元)', '结款情况', '已结清金额(万元)',
        '结算提成', '来源', '业主姓名', '业主电话', '服务内容', '备注', '编制人', '开始时间'
    ]

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, p in enumerate(projects, 2):
        data_row = [
            row_idx - 1,
            p.name or '',
            p.owner or '',
            p.location or '',
            p.total_investment,
            p.contract_amount,
            p.contract_status or '',
            p.invoice_status or '',
            p.invoiced_amount,
            p.payment_status or '',
            p.settled_amount,
            p.payment_settlement_status or '',
            p.source or '',
            p.owner_name or '',
            p.owner_phone or '',
            p.service_content or '',
            p.remark or '',
            p.author or '',
            p.start_date.strftime('%Y-%m-%d') if p.start_date else ''
        ]
        for col_idx, value in enumerate(data_row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    col_widths = [6, 28, 22, 18, 14, 14, 10, 10, 14, 10, 14, 10, 12, 10, 14, 20, 20, 10, 12]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'项目列表_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@projects_bp.route('/api/projects/search')
@login_required
def api_projects_search():
    context = request.args.get('context', 'index')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    keyword = request.args.get('keyword', '', type=str)
    year = request.args.get('year', 'all')
    month = request.args.get('month', 'all')
    settlement = request.args.get('settlement', 'all')
    start_date = request.args.get('start_date', '', type=str)
    end_date = request.args.get('end_date', '', type=str)
    contract_status = request.args.get('contract_status', '', type=str)
    payment_status = request.args.get('payment_status', '', type=str)
    payment_group = request.args.get('payment_group', '', type=str)
    invoice_status = request.args.get('invoice_status', '', type=str)

    query = Project.query.filter_by(is_valid=1)

    if context == 'my':
        if current_user.role == 'admin':
            pass
        elif current_user.role == 'visitor':
            query = query.filter(Project.visitors.any(User.id == current_user.id))
        else:
            query = query.filter_by(user_id=current_user.id)

    if context == 'admin' and current_user.role != 'admin':
        return jsonify({'success': False, 'message': '权限不足'}), 403

    if context == 'index':
        pass

    if keyword:
        from sqlalchemy import or_
        query = query.filter(or_(
            Project.name.contains(keyword),
            Project.owner.contains(keyword),
            Project.location.contains(keyword),
            Project.owner_name.contains(keyword),
            Project.service_content.contains(keyword),
            Project.remark.contains(keyword),
            Project.author.contains(keyword)
        ))

    if start_date:
        try:
            sd = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
            query = query.filter(Project.start_date >= sd)
        except ValueError:
            pass

    if end_date:
        try:
            ed = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(Project.start_date <= ed)
        except ValueError:
            pass

    if contract_status:
        query = query.filter(Project.contract_status == contract_status)

    if payment_group == 'received':
        query = query.filter(Project.payment_status.in_(['已结款', '已结清', '部分结清']))
    elif payment_group == 'receivable':
        query = query.filter(Project.payment_status.in_(['未结款', '未结', '未结算', '部分结清']))
    elif payment_status:
        query = query.filter(Project.payment_status == payment_status)

    if invoice_status:
        query = query.filter(Project.invoice_status == invoice_status)

    if year != 'all':
        query = query.filter(db.extract('year', Project.start_date) == int(year))

    if month != 'all':
        query = query.filter(db.extract('month', Project.start_date) == int(month))

    if settlement != 'all':
        query = query.filter(Project.payment_settlement_status == settlement)

    pagination = query.order_by(Project.updated_at.desc()).paginate(page=page, per_page=per_page, error_out=False)

    expense_map = {}
    if context == 'admin':
        from web.models import FundRecord
        from sqlalchemy import func
        expense_rows = db.session.query(
            FundRecord.project_id,
            func.sum(FundRecord.amount)
        ).filter(
            FundRecord.project_id.in_([p.id for p in pagination.items]),
            FundRecord.expense_type == '项目支出'
        ).group_by(FundRecord.project_id).all()
        expense_map = {int(r[0]): float(r[1] or 0) for r in expense_rows}

    # 查询附件数量
    project_ids = [p.id for p in pagination.items]
    attachment_counts = {}
    if project_ids:
        from web.models import Attachment
        attachments = Attachment.query.filter(Attachment.project_id.in_(project_ids)).all()
        for att in attachments:
            attachment_counts[att.project_id] = attachment_counts.get(att.project_id, 0) + 1

    rows = []
    for p in pagination.items:
        contract_amt = float(p.contract_amount or 0)
        settled_amt = float(p.settled_amount or 0)
        if p.payment_status in ('已结款', '已结清') and settled_amt == 0:
            settled_amt = contract_amt
        receivable_amt = max(0, contract_amt - settled_amt)

        rows.append({
            'id': p.id,
            'name': p.name,
            'owner': p.owner or '-',
            'total_investment': p.total_investment or '-',
            'contract_amount': p.contract_amount or '-',
            'contract_status': p.contract_status or '',
            'invoice_status': p.invoice_status or '',
            'invoiced_amount': p.invoiced_amount or 0,
            'payment_status': p.payment_status or '',
            'settled_amount': p.settled_amount or 0,
            'payment_settlement_status': p.payment_settlement_status or '',
            'source': p.source or '',
            'owner_name': p.owner_name or '',
            'owner_phone': p.owner_phone or '',
            'engineer_name': p.author or '未知',
            'start_date': p.start_date.strftime('%Y-%m-%d') if p.start_date else None,
            'start_date_full': p.start_date.strftime('%Y-%m-%d') if p.start_date else '',
            'location': p.location or '-',
            'received_amt': settled_amt,
            'receivable_amt': receivable_amt,
            'project_expense': expense_map.get(p.id, 0),
            'has_attachment': attachment_counts.get(p.id, 0) > 0
        })

    return jsonify({
        'success': True,
        'rows': rows,
        'pagination': {
            'page': pagination.page,
            'current_page': pagination.page,
            'per_page': pagination.per_page,
            'total': pagination.total,
            'total_pages': pagination.pages,
            'has_prev': pagination.has_prev,
            'has_next': pagination.has_next,
            'prev_num': pagination.prev_num if pagination.has_prev else 1,
            'next_num': pagination.next_num if pagination.has_next else pagination.page
        }
    })


@projects_bp.route('/project/detail/<int:project_id>')
@login_required
def project_detail(project_id):
    project = Project.query.get_or_404(project_id)

    if current_user.role == 'visitor':
        is_assigned = project.visitors.filter(User.id == current_user.id).count() > 0
        if not is_assigned:
            flash('您没有权限查看该项目')
            return redirect(url_for('projects.index'))

    attachments = Attachment.query.filter_by(project_id=project.id).order_by(Attachment.upload_time.desc()).all()
    logs = Log.query.filter_by(project_id=project.id).order_by(Log.time.desc()).all()

    is_owner = (current_user.id == project.user_id)
    can_edit = is_owner or (current_user.role == 'admin')
    can_delete = is_owner or (current_user.role == 'admin')

    return render_template('project_detail.html',
                           p=project,
                           attachments=attachments,
                           logs=logs,
                           can_delete=can_delete,
                           can_edit=can_edit,
                           is_owner=is_owner,
                           role=current_user.role)


@projects_bp.route('/project/add', methods=['GET', 'POST'])
@login_required
def project_add():
    if current_user.role == 'visitor':
        flash('访客没有新增项目权限')
        return redirect(url_for('projects.index'))

    if request.method == 'POST':
        name = request.form.get('name')
        description = request.form.get('description')
        location = request.form.get('location')
        project_type = request.form.get('project_type')
        phase = request.form.get('phase')
        start_date_str = request.form.get('start_date')
        start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else None

        project = Project(
            name=name,
            description=description,
            location=location,
            project_type=project_type,
            phase=phase,
            user_id=current_user.id,
            author=current_user.username,
            progress='新建',
            start_date=start_date,
            owner=request.form.get('owner'),
            total_investment=float(request.form.get('total_investment') or 0),
            contract_amount=float(request.form.get('contract_amount') or 0),
            contract_status=request.form.get('contract_status'),
            invoice_status=request.form.get('invoice_status'),
            invoiced_amount=float(request.form.get('invoiced_amount') or 0),
            payment_status=request.form.get('payment_status'),
            settled_amount=float(request.form.get('settled_amount') or 0),
            payment_settlement_status=request.form.get('payment_settlement_status', '未结提成'),
            source=request.form.get('source'),
            owner_name=request.form.get('owner_name'),
            owner_phone=request.form.get('owner_phone'),
            service_content=request.form.get('service_content'),
            remark=request.form.get('remark')
        )
        db.session.add(project)
        db.session.commit()

        log = Log(project_id=project.id, user=current_user.username, content='创建了项目')
        db.session.add(log)
        db.session.commit()

        from web.services.project_service import invalidate_project_stats_cache, find_duplicate_groups
        invalidate_project_stats_cache()

        duplicates = find_duplicate_groups(current_user.id)
        flash('项目创建成功')
        if duplicates:
            dup_count = sum(len(v) for v in duplicates.values()) - len(duplicates)
            flash(f'⚠ 检测到 {len(duplicates)} 组重复项目（共 {dup_count} 条重复记录），请检查核对', 'warning')
        return redirect(url_for('projects.my_projects'))


    return render_template('project_add.html', provinces=PROVINCES, role=current_user.role)


@projects_bp.route('/my_projects')
@login_required
def my_projects():
    filter_year = request.args.get('year', 'all')
    filter_month = request.args.get('month', 'all')
    filter_settlement = request.args.get('settlement', 'all')

    if current_user.role == 'admin':
        query = Project.query.filter_by(is_valid=1)
    elif current_user.role == 'visitor':
        query = Project.query.filter_by(is_valid=1).filter(
            Project.visitors.any(User.id == current_user.id)
        )
    else:
        query = Project.query.filter_by(user_id=current_user.id, is_valid=1)

    if filter_year != 'all':
        query = query.filter(db.extract('year', Project.start_date) == int(filter_year))

    if filter_month != 'all':
        query = query.filter(db.extract('month', Project.start_date) == int(filter_month))

    if filter_settlement != 'all':
        query = query.filter(Project.payment_settlement_status == filter_settlement)

    projects = query.order_by(Project.updated_at.desc()).all()

    # 查询每个项目是否有成果文件
    project_ids = [p.id for p in projects]
    attachment_counts = {}
    if project_ids:
        from web.models import Attachment
        attachments = Attachment.query.filter(Attachment.project_id.in_(project_ids)).all()
        for att in attachments:
            attachment_counts[att.project_id] = attachment_counts.get(att.project_id, 0) + 1

    now = datetime.datetime.now()
    years = list(range(now.year, now.year - 5, -1))
    months = list(range(1, 13))
    month_names = {1: '一月', 2: '二月', 3: '三月', 4: '四月', 5: '五月', 6: '六月',
                   7: '七月', 8: '八月', 9: '九月', 10: '十月', 11: '十一月', 12: '十二月'}

    from web.services.project_service import find_duplicate_groups
    if current_user.role == 'admin':
        dup_groups = find_duplicate_groups(None)
    elif current_user.role == 'visitor':
        dup_groups = {}
    else:
        dup_groups = find_duplicate_groups(current_user.id)

    dup_total_extra = sum(len(v) for v in dup_groups.values()) - len(dup_groups) if dup_groups else 0

    # 展示重复模式：当启用且存在重复时，只展示重复项目
    show_duplicates = request.args.get('show_duplicates', '0') == '1'
    if show_duplicates and dup_groups:
        dup_ids = set()
        for proj_list in dup_groups.values():
            for p in proj_list:
                dup_ids.add(p.id)
        projects = [p for p in projects if p.id in dup_ids]

    return render_template('my_projects.html',
                           projects=projects,
                           role=current_user.role,
                           years=years,
                           months=months,
                           month_names=month_names,
                           filter_year=filter_year,
                           filter_month=filter_month,
                           filter_settlement=filter_settlement,
                           dup_groups=dup_groups,
                           dup_total_extra=dup_total_extra,
                           show_duplicates=show_duplicates,
                           attachment_counts=attachment_counts)


@projects_bp.route('/my_projects/recheck', methods=['POST'])
@login_required
def my_projects_recheck():
    """用户修正完成后重新检测重复"""
    from web.services.project_service import find_duplicate_groups

    year = request.form.get('year', 'all')
    month = request.form.get('month', 'all')
    settlement = request.form.get('settlement', 'all')

    if current_user.role == 'admin':
        dup_groups = find_duplicate_groups(None)
    elif current_user.role != 'visitor':
        dup_groups = find_duplicate_groups(current_user.id)
    else:
        dup_groups = {}

    if dup_groups:
        flash(f'仍有 {len(dup_groups)} 组重复项目未处理，请继续修正', 'warning')
        return redirect(url_for('projects.my_projects', show_duplicates=1,
                                year=year, month=month, settlement=settlement))
    else:
        flash('所有重复项目已处理完毕，返回默认视图')
        return redirect(url_for('projects.my_projects'))


@projects_bp.route('/export_my_projects')
@login_required
def export_my_projects():
    filter_year = request.args.get('year', 'all')
    filter_month = request.args.get('month', 'all')
    filter_settlement = request.args.get('settlement', 'all')

    if current_user.role == 'admin':
        query = Project.query.filter_by(is_valid=1)
    elif current_user.role == 'visitor':
        query = Project.query.filter_by(is_valid=1).filter(
            Project.visitors.any(User.id == current_user.id)
        )
    else:
        query = Project.query.filter_by(user_id=current_user.id, is_valid=1)

    if filter_year != 'all':
        query = query.filter(db.extract('year', Project.start_date) == int(filter_year))

    if filter_month != 'all':
        query = query.filter(db.extract('month', Project.start_date) == int(filter_month))

    if filter_settlement != 'all':
        query = query.filter(Project.payment_settlement_status == filter_settlement)

    projects = query.order_by(Project.start_date.desc()).all()

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = '我的项目'

    headers = [
        '序号', '项目名称', '业主单位', '项目所在地', '总投资(万元)', '合同金额(万元)',
        '合同情况', '开票情况', '已开票金额(万元)', '结款情况', '已结清金额(万元)',
        '结算提成', '来源', '业主姓名', '业主电话', '服务内容', '备注', '编制人', '开始时间'
    ]

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, p in enumerate(projects, 2):
        data_row = [
            row_idx - 1,
            p.name or '',
            p.owner or '',
            p.location or '',
            p.total_investment,
            p.contract_amount,
            p.contract_status or '',
            p.invoice_status or '',
            p.invoiced_amount,
            p.payment_status or '',
            p.settled_amount,
            p.payment_settlement_status or '',
            p.source or '',
            p.owner_name or '',
            p.owner_phone or '',
            p.service_content or '',
            p.remark or '',
            p.author or '',
            p.start_date.strftime('%Y-%m-%d') if p.start_date else ''
        ]
        for col_idx, value in enumerate(data_row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    col_widths = [6, 28, 22, 18, 14, 14, 10, 10, 14, 10, 14, 10, 12, 10, 14, 20, 20, 10, 12]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'我的项目_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@projects_bp.route('/project/batch_upload', methods=['POST'])
@login_required
def batch_upload():
    if current_user.role == 'visitor':
        return jsonify({'success': False, 'message': '访客没有导入权限'}), 403

    file = request.files.get('file')
    if not file:
        return jsonify({'success': False, 'message': '未选择文件'}), 400

    filename = file.filename.lower()
    if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
        return jsonify({'success': False, 'message': '仅支持 .xlsx 或 .xls 格式'}), 400

    from openpyxl import load_workbook

    try:
        wb = load_workbook(file)
        ws = wb.active
    except Exception as e:
        return jsonify({'success': False, 'message': f'文件读取失败: {str(e)}'}), 400

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    if not rows:
        return jsonify({'success': False, 'message': '未找到数据行，请确保第一行为表头，第二行起为数据'}), 400

    COLUMN_MAP = {
        '项目名称': 'name',
        '业主单位': 'owner',
        '项目所在地': 'location',
        '总投资': 'total_investment',
        '合同金额': 'contract_amount',
        '合同情况': 'contract_status',
        '开票情况': 'invoice_status',
        '已开票金额': 'invoiced_amount',
        '结款情况': 'payment_status',
        '已结清金额': 'settled_amount',
        '结算提成': 'payment_settlement_status',
        '来源': 'source',
        '业主姓名': 'owner_name',
        '业主电话': 'owner_phone',
        '服务内容': 'service_content',
        '备注': 'remark',
        '开始时间': 'start_date',
        '项目开始时间': 'start_date',
    }

    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col_index = {}
    for idx, h in enumerate(header_row):
        if h and str(h).strip() in COLUMN_MAP:
            col_index[COLUMN_MAP[str(h).strip()]] = idx

    if 'name' not in col_index:
        return jsonify({'success': False, 'message': '表头中未找到"项目名称"列'}), 400

    success_count = 0
    errors = []

    for row_idx, row in enumerate(rows, 2):
        try:
            name = row[col_index['name']] if 'name' in col_index and row[col_index['name']] else None
            if not name or not str(name).strip():
                errors.append(f'第{row_idx}行: 项目名称为空')
                continue

            start_date = None
            date_field = col_index.get('start_date')
            if date_field is not None and row[date_field]:
                date_val = row[date_field]
                if isinstance(date_val, datetime.datetime):
                    start_date = date_val.date()
                elif isinstance(date_val, datetime.date):
                    start_date = date_val
                elif isinstance(date_val, str):
                    try:
                        start_date = datetime.datetime.strptime(date_val.strip(), '%Y-%m-%d').date()
                    except ValueError:
                        try:
                            start_date = datetime.datetime.strptime(date_val.strip(), '%Y/%m/%d').date()
                        except ValueError:
                            try:
                                start_date = datetime.datetime.strptime(date_val.strip(), '%Y年%m月%d日').date()
                            except ValueError:
                                errors.append(f'第{row_idx}行: 日期格式错误')
                                continue
                elif isinstance(date_val, (int, float)):
                    # Excel 序列号日期（如 46143 = 2026-05-01）
                    try:
                        start_date = (datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(date_val))).date()
                    except Exception:
                        errors.append(f'第{row_idx}行: 日期格式错误')
                        continue

            def get_float(field_name):
                idx = col_index.get(field_name)
                if idx is not None and idx < len(row) and row[idx] is not None:
                    try:
                        return float(row[idx])
                    except (ValueError, TypeError):
                        return 0.0
                return 0.0

            def get_str(field_name):
                idx = col_index.get(field_name)
                if idx is not None and idx < len(row) and row[idx] is not None:
                    return str(row[idx]).strip()
                return None

            project = Project(
                name=str(name).strip(),
                user_id=current_user.id,
                author=current_user.username,
                progress='新建',
                is_valid=1,
                start_date=start_date,
                owner=get_str('owner'),
                location=get_str('location'),
                total_investment=get_float('total_investment'),
                contract_amount=get_float('contract_amount'),
                contract_status=get_str('contract_status'),
                invoice_status=get_str('invoice_status'),
                invoiced_amount=get_float('invoiced_amount'),
                payment_status=get_str('payment_status'),
                settled_amount=get_float('settled_amount'),
                payment_settlement_status=get_str('payment_settlement_status') or '未结提成',
                source=get_str('source'),
                owner_name=get_str('owner_name'),
                owner_phone=get_str('owner_phone'),
                service_content=get_str('service_content'),
                remark=get_str('remark')
            )
            db.session.add(project)
            success_count += 1
        except Exception as e:
            errors.append(f'第{row_idx}行: {str(e)}')

    db.session.commit()

    from web.services.project_service import invalidate_project_stats_cache, find_duplicate_groups
    invalidate_project_stats_cache()

    duplicates = find_duplicate_groups(current_user.id)
    dup_warning = None
    if duplicates:
        dup_groups = []
        for (name, svc), projs in duplicates.items():
            dup_groups.append({
                'name': name,
                'service_content': svc,
                'project_ids': [p.id for p in projs],
                'project_names': [p.name for p in projs],
                'count': len(projs)
            })
        dup_warning = {
            'group_count': len(duplicates),
            'total_extra': sum(len(v) for v in duplicates.values()) - len(duplicates),
            'groups': dup_groups
        }

    message = f'成功导入 {success_count} 条项目'
    if errors:
        message += f'，{len(errors)} 条失败'

    result = {
        'success': True,
        'message': message,
        'errors': errors if errors else [],
        'count': success_count
    }
    if dup_warning:
        result['duplicates'] = dup_warning
        result['message'] += f'；检测到 {dup_warning["group_count"]} 组重复项目'
    return jsonify(result)


@projects_bp.route('/project/edit/<int:project_id>', methods=['GET', 'POST'])
@login_required
def project_edit(project_id):
    project = Project.query.get_or_404(project_id)

    if current_user.id != project.user_id and current_user.role != 'admin':
        flash('您没有权限编辑该项目')
        return redirect(url_for('projects.project_detail', project_id=project.id))

    if request.method == 'POST':
        from_dup = request.form.get('from_dup') == '1'
        project.name = request.form.get('name')
        project.description = request.form.get('description')
        project.location = request.form.get('location')
        project.project_type = request.form.get('project_type')
        project.phase = request.form.get('phase')
        project.updated_at = datetime.datetime.now()

        start_date_str = request.form.get('start_date')
        project.start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else None
        project.owner = request.form.get('owner')
        project.total_investment = float(request.form.get('total_investment') or 0)
        project.contract_amount = float(request.form.get('contract_amount') or 0)
        project.contract_status = request.form.get('contract_status')
        project.invoice_status = request.form.get('invoice_status')
        project.invoiced_amount = float(request.form.get('invoiced_amount') or 0)
        project.payment_status = request.form.get('payment_status')
        project.settled_amount = float(request.form.get('settled_amount') or 0)
        project.payment_settlement_status = request.form.get('payment_settlement_status')
        project.source = request.form.get('source')
        project.owner_name = request.form.get('owner_name')
        project.owner_phone = request.form.get('owner_phone')
        project.service_content = request.form.get('service_content')
        project.remark = request.form.get('remark')

        log = Log(project_id=project.id, user=current_user.username, content='编辑了项目信息')
        db.session.add(log)

        db.session.commit()
        flash('项目更新成功')
        if from_dup:
            return redirect(url_for('projects.my_projects', show_duplicates=1))
        return redirect(url_for('projects.project_detail', project_id=project.id))

    from_dup = request.args.get('from_dup') == '1'
    return render_template('project_edit.html', p=project, role=current_user.role, from_dup=from_dup)


@projects_bp.route('/project/update_settlement', methods=['POST'])
@login_required
def update_settlement():
    project_id = request.form.get('id')
    status = request.form.get('status')

    if not project_id:
        return jsonify({'success': False, 'message': '缺少项目ID'}), 400

    project = Project.query.get(int(project_id))
    if not project:
        return jsonify({'success': False, 'message': '项目不存在'}), 404

    if current_user.id != project.user_id and current_user.role != 'admin':
        return jsonify({'success': False, 'message': '没有权限'}), 403

    project.payment_settlement_status = status
    project.updated_at = datetime.datetime.now()

    log = Log(
        project_id=project.id,
        user=current_user.username,
        content=f'将结算提成状态更新为：{status}'
    )
    db.session.add(log)
    db.session.commit()

    return jsonify({'success': True, 'message': '更新成功'})


@projects_bp.route('/project/delete', methods=['POST'])
@login_required
def project_delete_by_form():
    project_id = request.form.get('id')
    redirect_dup = request.form.get('redirect_dup') == '1'
    if project_id:
        project = Project.query.get_or_404(int(project_id))
        if current_user.id == project.user_id or current_user.role == 'admin':
            project.is_valid = 0
            project.updated_at = datetime.datetime.now()
            db.session.commit()
            flash('项目已删除')
        else:
            flash('没有权限删除该项目')
    if redirect_dup:
        return redirect(url_for('projects.my_projects', show_duplicates=1))
    return redirect(url_for('projects.my_projects'))


@projects_bp.route('/project/delete/<int:project_id>')
@login_required
def project_delete(project_id):
    project = Project.query.get_or_404(project_id)
    if current_user.id == project.user_id or current_user.role == 'admin':
        project.is_valid = 0
        project.updated_at = datetime.datetime.now()
        db.session.commit()
        flash('项目已删除')
    else:
        flash('没有权限删除该项目')
    return redirect(url_for('projects.my_projects'))
