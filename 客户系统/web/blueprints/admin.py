import os
import io
import datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_from_directory, current_app, send_file
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from web.models import User, Project, FundRecord, StandardFile, Log
from web.extensions import db
from sqlalchemy import func, or_

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')


def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('需要管理员权限')
            return redirect(url_for('projects.index'))
        return f(*args, **kwargs)
    return decorated_function


@admin_bp.route('/users', methods=['GET', 'POST'])
@login_required
@admin_required
def users():
    if request.method == 'POST':
        name = request.form.get('name')
        role = request.form.get('role', 'engineer')
        password = request.form.get('password') or '123456'
        if not name:
            flash('用户名不能为空')
            return redirect(url_for('admin.users'))
        existing = User.query.filter_by(username=name).first()
        if existing:
            flash('用户名已存在')
            return redirect(url_for('admin.users'))
        user = User(username=name, role=role)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        if role == 'visitor':
            project_ids = request.form.getlist('visitor_projects')
            if project_ids:
                projects = Project.query.filter(Project.id.in_([int(pid) for pid in project_ids])).all()
                user.visitor_projects_rel = projects
                db.session.commit()
        flash('用户添加成功')
        return redirect(url_for('admin.users'))
    users_list = User.query.all()
    projects_list = Project.query.filter_by(is_valid=1).order_by(Project.name).all()
    return render_template('admin_users.html', users=users_list, projects=projects_list, role=current_user.role)


@admin_bp.route('/user/delete/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def user_delete(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash('不能删除自己')
        return redirect(url_for('admin.users'))
    if user.username == 'admin':
        flash('不能删除系统管理员')
        return redirect(url_for('admin.users'))
    project_count = Project.query.filter_by(user_id=user.id, is_valid=1).count()
    if project_count > 0:
        flash(f'该用户拥有 {project_count} 个项目，无法删除')
        return redirect(url_for('admin.users'))
    admin_user = User.query.filter_by(username='admin').first()
    if admin_user:
        Project.query.filter_by(user_id=user.id).update({'user_id': admin_user.id})
    db.session.delete(user)
    db.session.commit()
    flash('用户已删除')
    return redirect(url_for('admin.users'))


@admin_bp.route('/projects')
@login_required
@admin_required
def projects():
    page = request.args.get('page', 1, type=int)
    keyword = request.args.get('keyword', '', type=str)
    start_date = request.args.get('start_date', '', type=str)
    end_date = request.args.get('end_date', '', type=str)
    contract_status = request.args.get('contract_status', '', type=str)
    payment_status = request.args.get('payment_status', '', type=str)
    payment_group = request.args.get('payment_group', '', type=str)
    invoice_status = request.args.get('invoice_status', '', type=str)

    query = Project.query.filter_by(is_valid=1)

    if keyword:
        query = query.filter(or_(
            Project.name.contains(keyword),
            Project.owner.contains(keyword),
            Project.location.contains(keyword),
            Project.owner_name.contains(keyword),
            Project.service_content.contains(keyword),
            Project.remark.contains(keyword),
            Project.author.contains(keyword)
        ))
    if start_date:
        try:
            sd = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
            query = query.filter(Project.start_date >= sd)
        except ValueError:
            pass
    if end_date:
        try:
            ed = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(Project.start_date <= ed)
        except ValueError:
            pass
    if contract_status:
        query = query.filter(Project.contract_status == contract_status)
    if payment_group == 'received':
        query = query.filter(Project.payment_status.in_(['已结款', '已结清', '部分结清']))
    elif payment_group == 'receivable':
        query = query.filter(Project.payment_status.in_(['未结款', '未结', '未结算', '部分结清']))
    elif payment_status:
        query = query.filter(Project.payment_status == payment_status)
    if invoice_status:
        query = query.filter(Project.invoice_status == invoice_status)

    projects_paginated = query.order_by(Project.start_date.desc()).paginate(page=page, per_page=20, error_out=False)

    all_contract_statuses = sorted(set(
        s[0] for s in db.session.query(Project.contract_status).filter(Project.contract_status.isnot(None)).distinct()
    ))
    all_payment_statuses = sorted(set(
        s[0] for s in db.session.query(Project.payment_status).filter(Project.payment_status.isnot(None)).distinct()
    ))
    all_invoice_statuses = sorted(set(
        s[0] for s in db.session.query(Project.invoice_status).filter(Project.invoice_status.isnot(None)).distinct()
    ))

    payment_status_label = {
        'received': '已结款',
        'receivable': '未结款',
        '已结款': '已结款',
        '已结清': '已结清',
        '部分结清': '部分结清',
        '未结款': '未结款',
        '未结': '未结款',
        '未结算': '未结算'
    }.get(payment_status, payment_status)

    payment_group_label = ''
    if payment_group == 'received':
        payment_group_label = '已收账款（已结款 + 已结清 + 部分结清）'
    elif payment_group == 'receivable':
        payment_group_label = '应收账款（未结款 + 未结算 + 部分结清）'

    return render_template('admin_projects.html',
                           projects=projects_paginated,
                           keyword=keyword,
                           start_date=start_date,
                           end_date=end_date,
                           contract_status=contract_status,
                           payment_status=payment_status,
                           payment_group=payment_group,
                           invoice_status=invoice_status,
                           contract_statuses=all_contract_statuses,
                           payment_statuses=all_payment_statuses,
                           invoice_statuses=all_invoice_statuses,
                           payment_status_label=payment_status_label,
                           payment_group_label=payment_group_label,
                           role=current_user.role)


@admin_bp.route('/projects/export')
@login_required
@admin_required
def projects_export():
    keyword = request.args.get('keyword', '', type=str)
    start_date = request.args.get('start_date', '', type=str)
    end_date = request.args.get('end_date', '', type=str)
    contract_status = request.args.get('contract_status', '', type=str)
    payment_status = request.args.get('payment_status', '', type=str)
    payment_group = request.args.get('payment_group', '', type=str)
    invoice_status = request.args.get('invoice_status', '', type=str)

    query = Project.query.filter_by(is_valid=1)

    if keyword:
        query = query.filter(or_(
            Project.name.contains(keyword),
            Project.owner.contains(keyword),
            Project.location.contains(keyword),
            Project.owner_name.contains(keyword),
            Project.service_content.contains(keyword),
            Project.remark.contains(keyword),
            Project.author.contains(keyword)
        ))
    if start_date:
        try:
            sd = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
            query = query.filter(Project.start_date >= sd)
        except ValueError:
            pass
    if end_date:
        try:
            ed = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(Project.start_date <= ed)
        except ValueError:
            pass
    if contract_status:
        query = query.filter(Project.contract_status == contract_status)
    if payment_group == 'received':
        query = query.filter(Project.payment_status.in_(['已结款', '已结清', '部分结清']))
    elif payment_group == 'receivable':
        query = query.filter(Project.payment_status.in_(['未结款', '未结', '未结算', '部分结清']))
    elif payment_status:
        query = query.filter(Project.payment_status == payment_status)
    if invoice_status:
        query = query.filter(Project.invoice_status == invoice_status)

    projects_list = query.order_by(Project.start_date.desc()).all()

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = '项目列表'

    headers = [
        '序号', '项目名称', '业主单位', '项目所在地', '总投资(万元)', '合同金额(万元)',
        '合同情况', '开票情况', '已开票金额(万元)', '结款情况', '已结清金额(万元)',
        '结算提成', '来源', '业主姓名', '业主电话', '服务内容', '备注', '编制人', '开始时间'
    ]

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, p in enumerate(projects_list, 2):
        data_row = [
            row_idx - 1,
            p.name or '',
            p.owner or '',
            p.location or '',
            p.total_investment,
            p.contract_amount,
            p.contract_status or '',
            p.invoice_status or '',
            p.invoiced_amount,
            p.payment_status or '',
            p.settled_amount,
            p.payment_settlement_status or '',
            p.source or '',
            p.owner_name or '',
            p.owner_phone or '',
            p.service_content or '',
            p.remark or '',
            p.author or '',
            p.start_date.strftime('%Y-%m-%d') if p.start_date else ''
        ]
        for col_idx, value in enumerate(data_row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    col_widths = [6, 28, 22, 18, 14, 14, 10, 10, 14, 10, 14, 10, 12, 10, 14, 20, 20, 10, 12]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'项目列表_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@admin_bp.route('/project/delete/<int:project_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def project_delete(project_id):
    project = Project.query.get_or_404(project_id)
    Log.query.filter_by(project_id=project_id).delete()
    FundRecord.query.filter_by(project_id=project_id).delete()
    db.session.delete(project)
    db.session.commit()
    flash('项目已彻底删除')
    return redirect(url_for('admin.projects'))


@admin_bp.route('/operations', methods=['GET', 'POST'])
@login_required
@admin_required
def operations():
    if request.method == 'POST':
        amount = request.form.get('amount', type=float)
        expense_type = request.form.get('expense_type', '运营支出')
        project_id = request.form.get('project_id', type=int)
        use_date_str = request.form.get('use_date')
        purpose = request.form.get('purpose', '')
        remark = request.form.get('remark', '')

        if not amount or amount <= 0:
            flash('请输入有效金额')
            return redirect(url_for('admin.operations'))
        if not use_date_str:
            flash('请选择使用日期')
            return redirect(url_for('admin.operations'))

        try:
            use_date = datetime.datetime.strptime(use_date_str, '%Y-%m-%d').date()
        except ValueError:
            flash('日期格式无效')
            return redirect(url_for('admin.operations'))

        if expense_type == '项目支出' and not project_id:
            flash('项目支出必须选择关联项目')
            return redirect(url_for('admin.operations'))

        record = FundRecord(
            amount=amount,
            purpose=purpose or '未指定',
            remark=remark,
            use_date=use_date,
            create_time=datetime.datetime.utcnow(),
            create_user=current_user.username,
            expense_type=expense_type,
            project_id=project_id if expense_type == '项目支出' else None
        )
        db.session.add(record)
        db.session.commit()
        flash('支出记录添加成功')
        return redirect(url_for('admin.operations'))

    start_date = request.args.get('start_date', '', type=str)
    end_date = request.args.get('end_date', '', type=str)

    expense_query = FundRecord.query
    if start_date:
        try:
            sd = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
            expense_query = expense_query.filter(FundRecord.use_date >= sd)
        except ValueError:
            pass
    if end_date:
        try:
            ed = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
            expense_query = expense_query.filter(FundRecord.use_date <= ed)
        except ValueError:
            pass

    expense_list = expense_query.order_by(FundRecord.use_date.desc()).all()
    for e in expense_list:
        e.project_name = e.project.name if e.project else None

    total_expenses = sum(e.amount for e in expense_list) / 10000

    all_projects = Project.query.filter_by(is_valid=1).order_by(Project.name).all()

    all_valid_projects = Project.query.filter(Project.is_valid == 1).all()

    received = 0.0
    receivable = 0.0
    for p in all_valid_projects:
        contract = float(p.contract_amount or 0)
        settled = float(p.settled_amount or 0)
        if p.payment_status in ('已结款', '已结清') and settled == 0:
            settled = contract
        received += settled
        receivable += max(0, contract - settled)

    total_revenue = received + receivable

    project_expenses = sum(e.amount for e in expense_list if e.expense_type == '项目支出') / 10000
    operating_expenses = sum(e.amount for e in expense_list if e.expense_type == '运营支出') / 10000

    now = datetime.datetime.utcnow()
    month_labels = []
    month_revenues = []
    month_expenses_list = []
    month_profits = []

    for i in range(5, -1, -1):
        year = now.year
        month = now.month - i
        while month <= 0:
            month += 12
            year -= 1
        month_start = datetime.date(year, month, 1)
        if month == 12:
            month_end = datetime.date(year + 1, 1, 1)
        else:
            month_end = datetime.date(year, month + 1, 1)
        month_labels.append(month_start.strftime('%Y-%m'))

        month_revenue = 0.0
        for p in all_valid_projects:
            if p.start_date and month_start <= p.start_date < month_end:
                settled = float(p.settled_amount or 0)
                contract = float(p.contract_amount or 0)
                if p.payment_status in ('已结款', '已结清') and settled == 0:
                    settled = contract
                month_revenue += settled
        month_revenues.append(round(month_revenue, 2))

        month_expense = db.session.query(func.coalesce(func.sum(FundRecord.amount), 0)).filter(
            FundRecord.use_date >= month_start,
            FundRecord.use_date < month_end
        ).scalar()
        month_expense_wan = round(float(month_expense) / 10000, 2)
        month_expenses_list.append(month_expense_wan)
        month_profits.append(round(month_revenue - month_expense_wan, 2))

    return render_template('admin_operations.html',
                           expense_list=expense_list,
                           total_revenue=round(float(total_revenue or 0), 2),
                           total_expenses=round(total_expenses, 2),
                           received=round(received, 2),
                           receivable=round(receivable, 2),
                           project_expenses=round(project_expenses, 2),
                           operating_expenses=round(operating_expenses, 2),
                           month_labels=month_labels,
                           month_revenues=month_revenues,
                           month_expenses_list=month_expenses_list,
                           month_profits=month_profits,
                           all_projects=all_projects,
                           start_date=start_date,
                           end_date=end_date,
                           role=current_user.role)


@admin_bp.route('/operations/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def operations_delete(id):
    record = FundRecord.query.get_or_404(id)
    db.session.delete(record)
    db.session.commit()
    flash('支出记录已删除')
    return redirect(url_for('admin.operations'))


@admin_bp.route('/operations/export')
@login_required
@admin_required
def operations_export():
    start_date = request.args.get('start_date', '', type=str)
    end_date = request.args.get('end_date', '', type=str)

    expense_query = FundRecord.query
    if start_date:
        try:
            sd = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
            expense_query = expense_query.filter(FundRecord.use_date >= sd)
        except ValueError:
            pass
    if end_date:
        try:
            ed = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
            expense_query = expense_query.filter(FundRecord.use_date <= ed)
        except ValueError:
            pass

    expense_list = expense_query.order_by(FundRecord.use_date.desc()).all()

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = '支出记录'

    headers = ['序号', '使用日期', '金额(元)', '支出类型', '关联项目', '用途', '备注', '创建人', '创建时间']

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, e in enumerate(expense_list, 2):
        data_row = [
            row_idx - 1,
            e.use_date.strftime('%Y-%m-%d') if e.use_date else '',
            e.amount or 0,
            e.expense_type or '',
            e.project.name if e.project else '',
            e.purpose or '',
            e.remark or '',
            e.create_user or '',
            e.create_time.strftime('%Y-%m-%d %H:%M') if e.create_time else ''
        ]
        for col_idx, value in enumerate(data_row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    col_widths = [6, 14, 14, 12, 22, 22, 22, 12, 18]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'支出记录_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@admin_bp.route('/database')
@login_required
@admin_required
def database():
    file_type = request.args.get('file_type', '', type=str)
    upload_user = request.args.get('upload_user', '', type=str)

    query = StandardFile.query
    if file_type:
        query = query.filter(StandardFile.file_type == file_type)
    if upload_user:
        query = query.filter(StandardFile.upload_user == upload_user)

    files = query.order_by(StandardFile.upload_time.desc()).all()

    file_types = sorted(set(
        f[0] for f in db.session.query(StandardFile.file_type).filter(StandardFile.file_type.isnot(None)).distinct()
    ))
    upload_users = sorted(set(
        u[0] for u in db.session.query(StandardFile.upload_user).filter(StandardFile.upload_user.isnot(None)).distinct()
    ))

    return render_template('database.html',
                           files=files,
                           file_types=file_types,
                           upload_users=upload_users,
                           selected_file_type=file_type,
                           selected_upload_user=upload_user,
                           role=current_user.role)


@admin_bp.route('/database/upload', methods=['POST'])
@login_required
@admin_required
def database_upload():
    standard_name = request.form.get('standard_name', '')
    version = request.form.get('version', '1.0')
    file_type = request.form.get('file_type', '建设标准')
    file = request.files.get('file')

    if not file or file.filename == '':
        flash('请选择文件')
        return redirect(url_for('admin.database'))
    if not standard_name:
        flash('请输入标准名称')
        return redirect(url_for('admin.database'))

    standard_folder = current_app.config['STANDARD_FILES_FOLDER']
    os.makedirs(standard_folder, exist_ok=True)

    import uuid
    ext = os.path.splitext(file.filename)[1]
    save_name = f'{uuid.uuid4().hex}{ext}'
    file_path = os.path.join(standard_folder, save_name)
    file.save(file_path)

    standard_file = StandardFile(
        filename=file.filename,
        standard_name=standard_name,
        version=version,
        file_type=file_type,
        file_path=save_name,
        upload_time=datetime.datetime.utcnow(),
        upload_user=current_user.username
    )
    db.session.add(standard_file)
    db.session.commit()
    flash('文件上传成功')
    return redirect(url_for('admin.database'))


@admin_bp.route('/database/download/<int:file_id>')
@login_required
@admin_required
def database_download(file_id):
    sf = StandardFile.query.get_or_404(file_id)
    sf.download_count = (sf.download_count or 0) + 1
    db.session.commit()
    standard_folder = current_app.config['STANDARD_FILES_FOLDER']
    return send_from_directory(standard_folder, sf.file_path, as_attachment=True, download_name=sf.filename)


@admin_bp.route('/database/delete/<int:file_id>')
@login_required
@admin_required
def database_delete(file_id):
    sf = StandardFile.query.get_or_404(file_id)
    standard_folder = current_app.config['STANDARD_FILES_FOLDER']
    old_path = os.path.join(standard_folder, sf.file_path)
    if os.path.exists(old_path):
        os.remove(old_path)
    db.session.delete(sf)
    db.session.commit()
    flash('文件已删除')
    return redirect(url_for('admin.database'))


@admin_bp.route('/database/view/<int:file_id>')
@login_required
@admin_required
def database_view(file_id):
    sf = StandardFile.query.get_or_404(file_id)
    sf.download_count = (sf.download_count or 0) + 1
    db.session.commit()
    standard_folder = current_app.config['STANDARD_FILES_FOLDER']
    return send_from_directory(standard_folder, sf.file_path, download_name=sf.filename)


@admin_bp.route('/database/raw/<int:file_id>')
@login_required
@admin_required
def database_raw(file_id):
    sf = StandardFile.query.get_or_404(file_id)
    standard_folder = current_app.config['STANDARD_FILES_FOLDER']
    return send_from_directory(standard_folder, sf.file_path, as_attachment=False)
