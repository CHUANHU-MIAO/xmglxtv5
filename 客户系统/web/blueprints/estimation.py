from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app, send_file
from flask_login import login_required, current_user
from web.models import Project, EnergyFactor, InvestmentData
from web.extensions import db
import json
import math
import os
import io
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

estimation_bp = Blueprint('estimation', __name__)


def get_project_folder(project):
    return os.path.join(current_app.config['UPLOAD_FOLDER'], f"{project.id}-{project.name}")


def _check_visitor_access(project):
    if current_user.role == 'visitor':
        if not project.visitors.filter_by(id=current_user.id).first():
            flash('您没有权限访问该项目')
            return redirect(url_for('projects.index'))
    return None


def _ensure_project_folder(project):
    folder = get_project_folder(project)
    os.makedirs(folder, exist_ok=True)
    return folder


def _read_json(folder, filename):
    filepath = os.path.join(folder, filename)
    if os.path.exists(filepath):
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    return None


def _write_json(folder, filename, data):
    filepath = os.path.join(folder, filename)
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ============================================================
# 投资估算
# ============================================================

@estimation_bp.route('/project/<int:project_id>/investment')
@login_required
def investment_estimate(project_id):
    project = Project.query.get_or_404(project_id)
    check_result = _check_visitor_access(project)
    if check_result:
        return check_result
    return render_template('investment_estimate.html', project=project, role=current_user.role)


@estimation_bp.route('/api/project/<int:project_id>/investment/save', methods=['POST'])
@login_required
def investment_save(project_id):
    project = Project.query.get_or_404(project_id)
    check_result = _check_visitor_access(project)
    if check_result:
        return jsonify({'success': False, 'message': '无权限'})

    data = request.json.get('data', [])

    folder = _ensure_project_folder(project)
    _write_json(folder, 'investment_data.json', data)

    try:
        InvestmentData.query.filter_by(project_id=project_id).delete()

        for item in data:
            index_val = item.get('index', 0)
            try:
                index_val = float(index_val)
            except (ValueError, TypeError):
                index_val = 0

            inv = InvestmentData(
                project_id=project_id,
                serial_number=item.get('serial_number', ''),
                item_name=item.get('item_name', ''),
                building_cost=item.get('building_cost', 0),
                installation_cost=item.get('installation_cost', 0),
                equipment_cost=item.get('equipment_cost', 0),
                other_cost=item.get('other_cost', 0),
                unit=item.get('unit', ''),
                quantity=item.get('quantity', 0),
                index=index_val,
                use_index=item.get('use_index', False),
                billing_basis=item.get('billing_basis', ''),
                calc_rate=item.get('calc_rate', 0),
                discount_rate=item.get('discount_rate', 100),
                build_category=item.get('build_category', ''),
                address_category=item.get('address_category', ''),
                is_reserve_rate=item.get('is_reserve_rate', False),
                reserve_rate=item.get('reserve_rate')
            )
            db.session.add(inv)

        db.session.commit()
        return jsonify({'success': True, 'message': '保存成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})


@estimation_bp.route('/api/project/<int:project_id>/investment/load')
@login_required
def investment_load(project_id):
    project = Project.query.get_or_404(project_id)
    check_result = _check_visitor_access(project)
    if check_result:
        return jsonify({'success': False, 'message': '无权限'})

    folder = get_project_folder(project)
    json_data = _read_json(folder, 'investment_data.json')

    if json_data is not None:
        return jsonify({'success': True, 'data': json_data})

    items = InvestmentData.query.filter_by(project_id=project_id).order_by(InvestmentData.serial_number).all()

    data = []
    for item in items:
        data.append({
            'serial_number': item.serial_number,
            'item_name': item.item_name,
            'building_cost': item.building_cost,
            'installation_cost': item.installation_cost,
            'equipment_cost': item.equipment_cost,
            'other_cost': item.other_cost,
            'unit': item.unit,
            'quantity': item.quantity,
            'index': item.index,
            'use_index': item.use_index,
            'billing_basis': item.billing_basis,
            'calc_rate': item.calc_rate,
            'discount_rate': item.discount_rate,
            'build_category': item.build_category,
            'address_category': item.address_category,
            'is_reserve_rate': item.is_reserve_rate,
            'reserve_rate': item.reserve_rate
        })

    return jsonify({'success': True, 'data': data})


@estimation_bp.route('/api/project/<int:project_id>/investment/export')
@login_required
def investment_export(project_id):
    project = Project.query.get_or_404(project_id)
    check_result = _check_visitor_access(project)
    if check_result:
        return jsonify({'success': False, 'message': '无权限'})

    folder = get_project_folder(project)
    json_data = _read_json(folder, 'investment_data.json')
    data = json_data if json_data is not None else []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '投资估算'

    headers = ['序号', '工程或费用名称', '建筑工程费', '安装工程费', '设备购置与安装费', '其他费用', '合计', '单位', '数量', '单价']
    ws.append(headers)

    for item in data:
        serial = item.get('serial_number', '')
        is_level2 = serial.startswith('2.') and len(serial.split('.')) == 2
        building = item.get('building_cost', 0)
        installation = item.get('installation_cost', 0)
        equipment = item.get('equipment_cost', 0)
        other = item.get('other_cost', 0)
        total = item.get('total', building + installation + equipment + other)

        if is_level2:
            col7 = item.get('billing_basis', '')
            col8 = item.get('calc_rate', 0)
        else:
            col7 = item.get('unit', '')
            col8 = item.get('quantity', 0)

        col9 = item.get('index', 0)

        row_data = [
            serial,
            item.get('item_name', ''),
            building,
            installation,
            equipment,
            other,
            total,
            col7 if col7 is not None else '',
            col8 if col8 is not None else '',
            col9 if col9 is not None else ''
        ]
        ws.append(row_data)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell_font = Font(name='宋体', size=10.5)
    cell_alignment = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                cell_len = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                max_length = max(max_length, cell_len)
        ws.column_dimensions[col_letter].width = max_length + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{project.name}_投资估算.xlsx"
    if current_user.role == 'visitor':
        from datetime import datetime as dt
        current_user.last_active_time = dt.now()
        db.session.commit()
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================================
# 能耗估算
# ============================================================

@estimation_bp.route('/project/<int:project_id>/energy')
@login_required
def energy_estimate(project_id):
    project = Project.query.get_or_404(project_id)
    check_result = _check_visitor_access(project)
    if check_result:
        return check_result
    return render_template('energy_estimate.html', project=project, role=current_user.role)


@estimation_bp.route('/api/energy/factors')
@login_required
def energy_factors():
    factors = EnergyFactor.query.filter_by(is_active=True).order_by(EnergyFactor.sort_order, EnergyFactor.id).all()
    data = []
    for f in factors:
        data.append({
            'id': f.id,
            'name': f.name,
            'unit': f.unit,
            'equivalent_coef': f.equivalent_coef,
            'equivalent_note': f.equivalent_note,
            'equivalent_coef_val': f.equivalent_coef_val,
            'equivalent_val_note': f.equivalent_val_note,
            'category': f.category
        })
    return jsonify({'success': True, 'factors': data})


@estimation_bp.route('/api/project/<int:project_id>/energy/params', methods=['GET', 'POST'])
@login_required
def energy_params(project_id):
    project = Project.query.get_or_404(project_id)
    check_result = _check_visitor_access(project)
    if check_result:
        if request.method == 'POST':
            return jsonify({'success': False, 'message': '无权限'})
        return jsonify({'success': False, 'message': '无权限'})

    folder = _ensure_project_folder(project)

    if request.method == 'GET':
        data = _read_json(folder, 'energy_params.json')
        return jsonify({'success': True, 'data': data if data is not None else {}})

    if request.method == 'POST':
        data = request.json
        _write_json(folder, 'energy_params.json', data)
        return jsonify({'success': True, 'message': '保存成功'})


@estimation_bp.route('/api/project/<int:project_id>/energy/items', methods=['GET', 'POST'])
@login_required
def energy_items(project_id):
    project = Project.query.get_or_404(project_id)
    check_result = _check_visitor_access(project)
    if check_result:
        if request.method == 'POST':
            return jsonify({'success': False, 'message': '无权限'})
        return jsonify({'success': False, 'message': '无权限'})

    folder = _ensure_project_folder(project)

    if request.method == 'GET':
        data = _read_json(folder, 'energy_items.json')
        return jsonify({'success': True, 'data': data if data is not None else []})

    if request.method == 'POST':
        data = request.json
        _write_json(folder, 'energy_items.json', data)
        return jsonify({'success': True, 'message': '保存成功'})


@estimation_bp.route('/api/project/<int:project_id>/energy/electricity', methods=['GET', 'POST'])
@login_required
def energy_electricity(project_id):
    project = Project.query.get_or_404(project_id)
    check_result = _check_visitor_access(project)
    if check_result:
        if request.method == 'POST':
            return jsonify({'success': False, 'message': '无权限'})
        return jsonify({'success': False, 'message': '无权限'})

    folder = _ensure_project_folder(project)

    if request.method == 'GET':
        data = _read_json(folder, 'energy_electricity.json')
        return jsonify({'success': True, 'data': data if data is not None else {}})

    if request.method == 'POST':
        data = request.json
        _write_json(folder, 'energy_electricity.json', data)
        return jsonify({'success': True, 'message': '保存成功'})


@estimation_bp.route('/api/project/<int:project_id>/energy/export')
@login_required
def energy_export(project_id):
    project = Project.query.get_or_404(project_id)
    check_result = _check_visitor_access(project)
    if check_result:
        return jsonify({'success': False, 'message': '无权限'})

    try:
        folder = get_project_folder(project)
        items_path = os.path.join(folder, 'energy_items.json')
        elec_path = os.path.join(folder, 'energy_electricity.json')

        items = []
        if os.path.exists(items_path):
            with open(items_path, 'r', encoding='utf-8') as f:
                items = json.load(f)

        elec_data = {'rows': [], 'simultaneous': {'ksp': 0.90, 'ksq': 0.93}, 'compensation_kvar': 0}
        if os.path.exists(elec_path):
            with open(elec_path, 'r', encoding='utf-8') as f:
                elec_data = json.load(f)

        wb = openpyxl.Workbook()
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell_font = Font(name='宋体', size=10.5)
        cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_font = Font(name='宋体', size=10.5, bold=True)

        ws = wb.active
        ws.title = '能耗估算汇总'

        ws.append(['用电负荷计算表'])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
        ws.cell(row=1, column=1).font = Font(name='宋体', size=14, bold=True)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

        elec_headers = ['序号', '负荷名称', '电压(kV)', '用电密度值(W/m²)', '用电数量(m²)',
                         '设备容量工作(kW)', '设备容量总计(kW)', '需用系数KC', 'cosφ', 'tgφ',
                         '有功功率(kW)', '无功功率(kvar)', '视在功率(kVA)',
                         '年利用小时数', '电耗(kW·h)', '备注']
        ws.append(elec_headers)
        header_row = ws.max_row
        for col_idx, h in enumerate(elec_headers, 1):
            c = ws.cell(row=header_row, column=col_idx)
            c.font = header_font
            c.alignment = cell_alignment
            c.border = thin_border

        elec_rows = elec_data.get('rows', [])
        for i, row in enumerate(elec_rows, 1):
            density = float(row.get('density', 0))
            density_qty = float(row.get('density_qty', 0))
            work_kw = float(row.get('work_kw', 0)) or density * density_qty / 1000
            total_kw = float(row.get('total_kw', 0)) or work_kw
            kc = float(row.get('kc', 0))
            cos_phi = float(row.get('cos_phi', 0.90))
            tg_phi = round(math.sin(math.acos(min(cos_phi, 0.999))) / cos_phi if cos_phi > 0 else 0, 4)
            active_kw = round(total_kw * kc, 2)
            reactive_kvar = round(active_kw * tg_phi, 2)
            apparent_kva = round(active_kw / cos_phi if cos_phi > 0 else 0, 2)
            annual_hours = float(row.get('annual_hours', 0))
            power_consumption = round(active_kw * annual_hours, 2)

            vals = [i, row.get('name', ''), row.get('voltage', 0.38),
                    density, density_qty, round(work_kw, 2), round(total_kw, 2),
                    kc, cos_phi, tg_phi, active_kw, reactive_kvar, apparent_kva,
                    annual_hours, power_consumption, row.get('remark', '')]
            ws.append(vals)
            for col_idx in range(1, len(vals) + 1):
                ws.cell(row=ws.max_row, column=col_idx).font = cell_font
                ws.cell(row=ws.max_row, column=col_idx).alignment = cell_alignment
                ws.cell(row=ws.max_row, column=col_idx).border = thin_border

        sum_work = sum(float(r.get('work_kw', 0)) or float(r.get('density', 0)) * float(r.get('density_qty', 0)) / 1000 for r in elec_rows)
        sum_total = sum(float(r.get('total_kw', 0)) or (float(r.get('density', 0)) * float(r.get('density_qty', 0)) / 1000) for r in elec_rows)
        sum_active = 0; sum_reactive = 0; sum_apparent = 0; sum_consumption = 0
        for r in elec_rows:
            tk = float(r.get('total_kw', 0)) or float(r.get('work_kw', 0)) or float(r.get('density', 0)) * float(r.get('density_qty', 0)) / 1000
            kc = float(r.get('kc', 0))
            cp = float(r.get('cos_phi', 0.90))
            tg = math.sin(math.acos(min(cp, 0.999))) / cp if cp > 0 else 0
            ak = tk * kc
            sum_active += ak
            sum_reactive += ak * tg
            sum_apparent += ak / cp if cp > 0 else 0
            sum_consumption += ak * float(r.get('annual_hours', 0))
        ksp = float(elec_data.get('simultaneous', {}).get('ksp', 0.90))
        ksq = float(elec_data.get('simultaneous', {}).get('ksq', 0.93))
        comp = float(elec_data.get('compensation_kvar', 0))
        sim_active = round(sum_active * ksp, 2)
        sim_reactive = round(sum_reactive * ksq, 2)
        sim_apparent = round(math.sqrt(sim_active * sim_active + sim_reactive * sim_reactive), 2)
        final_reactive = round(max(0, sim_reactive - comp), 2)
        final_apparent = round(math.sqrt(sim_active * sim_active + final_reactive * final_reactive), 2)

        total_row = ws.max_row + 1
        ws.append(['合计', '', '', '', '', round(sum_work, 2), round(sum_total, 2), '', '', '',
                   round(sum_active, 2), round(sum_reactive, 2), round(sum_apparent, 2), '', round(sum_consumption, 2), ''])
        for col_idx in range(1, 17):
            ws.cell(row=total_row, column=col_idx).font = Font(name='宋体', size=10.5, bold=True)
            ws.cell(row=total_row, column=col_idx).alignment = cell_alignment
            ws.cell(row=total_row, column=col_idx).border = thin_border

        sim_row = ws.max_row + 1
        ws.append(['乘以同时系数', '', '', '', '', '', '', f'Kp={ksp}', '', f'Kq={ksq}',
                   sim_active, sim_reactive, sim_apparent, '', '', ''])
        for col_idx in range(1, 17):
            ws.cell(row=sim_row, column=col_idx).font = Font(name='宋体', size=10.5, bold=True)
            ws.cell(row=sim_row, column=col_idx).alignment = cell_alignment
            ws.cell(row=sim_row, column=col_idx).border = thin_border

        comp_row = ws.max_row + 1
        ws.append(['无功补偿', '', '', '', '', '', '', '', '', '', '', f'{comp} kvar', '', '', '', ''])
        for col_idx in range(1, 17):
            ws.cell(row=comp_row, column=col_idx).font = Font(name='宋体', size=10.5, bold=True)
            ws.cell(row=comp_row, column=col_idx).alignment = cell_alignment
            ws.cell(row=comp_row, column=col_idx).border = thin_border

        final_row = ws.max_row + 1
        ws.append(['补偿后', '', '', '', '', '', '', '', '', '', sim_active, final_reactive, final_apparent, '', '', ''])
        for col_idx in range(1, 17):
            ws.cell(row=final_row, column=col_idx).font = Font(name='宋体', size=10.5, bold=True)
            ws.cell(row=final_row, column=col_idx).alignment = cell_alignment
            ws.cell(row=final_row, column=col_idx).border = thin_border

        for _ in range(4):
            ws.append([])

        energy_start_row = ws.max_row + 1
        ws.append(['项目能源消耗估算表'])
        ws.merge_cells(start_row=energy_start_row, start_column=1, end_row=energy_start_row, end_column=11)
        ws.cell(row=energy_start_row, column=1).font = Font(name='宋体', size=14, bold=True)
        ws.cell(row=energy_start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')

        energy_headers = ['序号', '能源/耗能工质名称', '计量单位', '年消耗量',
                           '折标煤系数(当量)', '当量值标准煤(tce)',
                           '折标煤系数(等价)', '等价值标准煤(tce)',
                           '单价(含税)', '年费用(万元)', '消耗环节/备注']
        ws.append(energy_headers)
        eh_row = ws.max_row
        for col_idx, h in enumerate(energy_headers, 1):
            c = ws.cell(row=eh_row, column=col_idx)
            c.font = header_font
            c.alignment = cell_alignment
            c.border = thin_border

        for i, item in enumerate(items, 1):
            annual_qty = float(item.get('annual_qty', 0))
            eq_coef = float(item.get('equivalent_coef', 0))
            ev_coef = float(item.get('equivalent_coef_val', 0))
            unit_price = float(item.get('unit_price', 0))
            eq_tce = annual_qty * eq_coef
            ev_tce = annual_qty * ev_coef
            annual_cost = annual_qty * unit_price / 10000
            vals = [i, item.get('name', ''), item.get('unit', ''),
                    annual_qty, eq_coef, round(eq_tce, 2),
                    ev_coef, round(ev_tce, 2),
                    unit_price, round(annual_cost, 2),
                    item.get('remark', '')]
            ws.append(vals)
            for col_idx in range(1, len(vals) + 1):
                ws.cell(row=ws.max_row, column=col_idx).font = cell_font
                ws.cell(row=ws.max_row, column=col_idx).alignment = cell_alignment
                ws.cell(row=ws.max_row, column=col_idx).border = thin_border

        total_eq = sum(float(it.get('annual_qty', 0)) * float(it.get('equivalent_coef', 0)) for it in items)
        total_ev = sum(float(it.get('annual_qty', 0)) * float(it.get('equivalent_coef_val', 0)) for it in items)
        total_cost = sum(float(it.get('annual_qty', 0)) * float(it.get('unit_price', 0)) / 10000 for it in items)
        ttl_row = ws.max_row + 1
        ws.append(['合计', '', '', '', '', round(total_eq, 2), '', round(total_ev, 2), '', round(total_cost, 2), ''])
        for col_idx in range(1, 12):
            ws.cell(row=ttl_row, column=col_idx).font = Font(name='宋体', size=10.5, bold=True)
            ws.cell(row=ttl_row, column=col_idx).alignment = cell_alignment
            ws.cell(row=ttl_row, column=col_idx).border = thin_border

        for col in ws.columns:
            max_length = 0
            try:
                col_letter = col[0].column_letter
            except AttributeError:
                continue
            for cell in col:
                try:
                    if cell.value:
                        cell_len = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                        max_length = max(max_length, cell_len)
                except AttributeError:
                    continue
            ws.column_dimensions[col_letter].width = min(max_length + 6, 24)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"{project.name}_能耗估算汇总.xlsx"
        if current_user.role == 'visitor':
            from datetime import datetime as dt
            current_user.last_active_time = dt.now()
            db.session.commit()
        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        flash(f'导出失败: {str(e)}')
        return redirect(f'/project/{project_id}/energy')


# ============================================================
# 财务分析
# ============================================================

@estimation_bp.route('/project/<int:project_id>/finance')
@login_required
def finance_select(project_id):
    project = Project.query.get_or_404(project_id)
    check_result = _check_visitor_access(project)
    if check_result:
        return check_result
    return render_template('finance_select.html', project=project, role=current_user.role)


@estimation_bp.route('/project/<int:project_id>/finance/bond')
@login_required
def finance_bond(project_id):
    project = Project.query.get_or_404(project_id)
    check_result = _check_visitor_access(project)
    if check_result:
        return check_result
    return render_template('finance_bond.html', project=project, role=current_user.role)


@estimation_bp.route('/project/<int:project_id>/finance/normal')
@login_required
def finance_normal(project_id):
    project = Project.query.get_or_404(project_id)
    check_result = _check_visitor_access(project)
    if check_result:
        return check_result
    return render_template('finance_normal.html', project=project, role=current_user.role)


@estimation_bp.route('/api/project/<int:project_id>/finance/save', methods=['POST'])
@login_required
def finance_save(project_id):
    project = Project.query.get_or_404(project_id)
    check_result = _check_visitor_access(project)
    if check_result:
        return jsonify({'success': False, 'message': '无权限'})

    data = request.json or {}
    mode = data.get('mode', 'normal')
    payload = data.get('data', {})

    folder = _ensure_project_folder(project)

    if mode == 'bond':
        _write_json(folder, 'finance_bond.json', payload)
    else:
        _write_json(folder, 'finance_normal.json', payload)

    return jsonify({'success': True, 'message': '保存成功'})


@estimation_bp.route('/api/project/<int:project_id>/finance/load')
@login_required
def finance_load(project_id):
    project = Project.query.get_or_404(project_id)
    check_result = _check_visitor_access(project)
    if check_result:
        return jsonify({'success': False, 'message': '无权限'})

    mode = request.args.get('mode', 'normal')
    folder = get_project_folder(project)

    if mode == 'bond':
        data = _read_json(folder, 'finance_bond.json')
    else:
        data = _read_json(folder, 'finance_normal.json')

    return jsonify({'success': True, 'data': data if data is not None else {}})


@estimation_bp.route('/api/project/<int:project_id>/finance/meta')
@login_required
def finance_meta(project_id):
    project = Project.query.get_or_404(project_id)
    check_result = _check_visitor_access(project)
    if check_result:
        return jsonify({'success': False, 'message': '无权限'})

    folder = get_project_folder(project)
    data = _read_json(folder, 'finance_meta.json')
    return jsonify({'success': True, 'data': data if data is not None else {}})


@estimation_bp.route('/api/project/<int:project_id>/finance/export')
@login_required
def finance_export(project_id):
    project = Project.query.get_or_404(project_id)
    check_result = _check_visitor_access(project)
    if check_result:
        return jsonify({'success': False, 'message': '无权限'})

    mode = request.args.get('mode', 'bond')
    folder = get_project_folder(project)

    data = {}
    if mode == 'bond':
        data = _read_json(folder, 'finance_bond.json') or {}
    else:
        data = _read_json(folder, 'finance_normal.json') or {}

    try:
        wb = openpyxl.Workbook()
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell_font = Font(name='宋体', size=10.5)
        cell_alignment = Alignment(horizontal='center', vertical='center')
        header_font = Font(name='宋体', size=10.5, bold=True)

        fund = data.get('fund', {})
        repay = data.get('repay', {})
        rc = data.get('revcost', {})
        tax = data.get('tax', {})

        ws1 = wb.active
        ws1.title = '资金使用计划表'
        ws1.append(['资金来源', '金额（万元）', '占比', '备注'])
        total_inv = fund.get('totalInvestment', 0)
        fund_items = [('估算总投资', fund.get('totalInvestment', 0)),
                     ('一、资本金', fund.get('ownCapital', 0)),
                     ('\u3000自有资金', fund.get('selfFund', 0)),
                     ('二、债务资金', fund.get('debtFund', 0))]
        if mode == 'bond':
            fund_items.append(('\u3000专项债券', fund.get('bondFund', 0)))
        fund_items.append(('\u3000银行融资', fund.get('bankLoan', 0)))
        for item in fund_items:
            pct = item[1] / total_inv if total_inv > 0 else 0
            ws1.append([item[0], item[1], f'{pct:.2%}', ''])

        ws2 = wb.create_sheet('还本付息')
        debt_list = repay.get('debtList', [])
        if debt_list:
            ws2.append(['债务名称', '起始年', '期限(年)', '金额(万元)', '利率(%)', '还本方式'])
            for d in debt_list:
                method_map = {'equal-principal': '等额本金', 'equal-installment': '等额本息', 'lump-sum': '到期一次性还本'}
                ws2.append([d.get('name', ''), d.get('startYear', ''), d.get('duration', ''), d.get('amount', 0), d.get('rate', 0), method_map.get(d.get('method', ''), d.get('method', ''))])
        else:
            ws2.append(['暂无债务数据'])

        ws3 = wb.create_sheet('收入成本')
        ws3.append(['收入项目', '数量', '单价(万元/年)', '使用率(%)', '增长率(%)'])
        for r in rc.get('revenueItems', rc.get('revenues', [])):
            ws3.append([r.get('name', ''), r.get('qty', r.get('spaces', 0)), r.get('price', 0), r.get('rate', 100), r.get('growth', 0)])
        ws3.append([])
        ws3.append(['成本项目', '金额（万元/年）', '增长率(%)'])
        for c in rc.get('costItems', rc.get('costs', [])):
            ws3.append([c.get('name', ''), c.get('amount', 0), c.get('growth', 0)])

        ws4 = wb.create_sheet('税费参数')
        ws4.append(['参数', '值'])
        ws4.append(['销项税率(%)', tax.get('taxRateOutput', 0)])
        ws4.append(['进项税率(%)', tax.get('taxRateInput', 0)])
        ws4.append(['城建税税率(%)', tax.get('taxRateCity', 0)])
        ws4.append(['教育费附加税率(%)', tax.get('taxRateEdu', 0)])
        ws4.append(['地方教育费附加税率(%)', tax.get('taxRateLocalEdu', 0)])
        ws4.append(['建设期增值税抵扣', '是' if tax.get('vatDeduction', False) else '否'])

        sheets = [ws1, ws2, ws3, ws4]

        if mode == 'normal':
            nb = data.get('nb', {})
            ws5 = wb.create_sheet('折旧摊销参数')
            ws5.append(['参数', '值'])
            ws5.append(['房屋折旧年限(年)', nb.get('nbDepYears', 20)])
            ws5.append(['房屋残值率(%)', nb.get('nbDepResidualRate', 5)])
            ws5.append(['设备折旧年限(年)', nb.get('nbEquipDepYears', 15)])
            ws5.append(['设备残值率(%)', nb.get('nbEquipResidualRate', 5)])
            ws5.append(['摊销年限(年)', nb.get('nbAmortYears', 40)])
            ws5.append(['基准收益率(%)', nb.get('nbDiscountRate', 8)])
            sheets.append(ws5)

            ws6 = wb.create_sheet('项目基本信息')
            ws6.append(['参数', '值'])
            ws6.append(['项目名称', fund.get('projectName', '')])
            ws6.append(['建设期起始年', fund.get('constructionStartYear', '')])
            ws6.append(['建设期结束年', fund.get('constructionEndYear', '')])
            ws6.append(['运营期起始年', fund.get('operationStartYear', '')])
            ws6.append(['运营期结束年', fund.get('operationEndYear', '')])
            cs = fund.get('constructionStartYear', 0)
            oe = fund.get('operationEndYear', 0)
            ws6.append(['计算周期(年)', oe - cs + 1 if cs and oe else ''])
            ws6.append(['估算总投资(万元)', total_inv])
            ws6.append(['折旧年限(年)', rc.get('depreciationYears', 20)])
            ws6.append(['残值率(%)', rc.get('residualRate', 5)])
            ws6.append(['摊销年限(年)', rc.get('amortizationYears', 20)])
            sheets.append(ws6)

        for ws in sheets:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.font = cell_font
                    cell.alignment = cell_alignment
                    cell.border = thin_border
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        cell_len = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                        max_length = max(max_length, cell_len)
                ws.column_dimensions[col_letter].width = max_length + 4

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        mode_label = '专项债' if mode == 'bond' else '非专项债'
        filename = f"{project.name}_财务测算_{mode_label}.xlsx"

        if current_user.role == 'visitor':
            from datetime import datetime as dt
            current_user.last_active_time = dt.now()
            db.session.commit()

        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        flash(f'导出失败: {str(e)}')
        target = 'bond' if mode == 'bond' else 'normal'
        return redirect(f'/project/{project_id}/finance/{target}')


@estimation_bp.route('/api/project/<int:project_id>/finance/export-all', methods=['POST'])
@login_required
def finance_export_all(project_id):
    project = Project.query.get_or_404(project_id)
    check_result = _check_visitor_access(project)
    if check_result:
        return jsonify({'success': False, 'message': '无权限'})

    try:
        data = request.get_json()
        mode = data.get('mode', 'bond')
        sheets_data = data.get('sheets', [])

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell_font = Font(name='宋体', size=10.5)
        cell_alignment = Alignment(horizontal='center', vertical='center')

        def try_float(v):
            if isinstance(v, (int, float)):
                return v
            if isinstance(v, str):
                try:
                    return float(v.replace(',', ''))
                except ValueError:
                    return v
            return v

        for sd in sheets_data:
            name = sd.get('name', 'Sheet')[:31]
            ws = wb.create_sheet(title=name)
            tables = sd.get('tables', [])
            summary = sd.get('summary', [])

            for ti, table in enumerate(tables):
                headers = table.get('headers', [])
                rows = table.get('rows', [])
                if headers:
                    ws.append(headers)
                for row in rows:
                    ws.append([try_float(c) for c in row])
                if ti < len(tables) - 1 or summary:
                    for _ in range(5):
                        ws.append([])

            if summary:
                for item in summary:
                    ws.append([try_float(item[0]), try_float(item[1])])

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.font = cell_font
                    cell.alignment = cell_alignment
                    cell.border = thin_border

            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        cell_len = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                        max_length = max(max_length, cell_len)
                ws.column_dimensions[col_letter].width = max_length + 4

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        mode_label = '专项债' if mode == 'bond' else '一般项目'
        filename = f"{project.name}_财务测算_{mode_label}.xlsx"
        if current_user.role == 'visitor':
            from datetime import datetime as dt
            current_user.last_active_time = dt.now()
            db.session.commit()
        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
